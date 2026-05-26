"""
BOQ Excel Generator — generate 6-sheet Bill of Quantities Excel from JSON.

Phase 3 — Task #20B · Chairman approved 2026-05-26.

Follows TT 13/2021/TT-BXD format (replaces QĐ 1129/QĐ-BXD):
  Sheet 1: Tổng hợp        (Summary — totals + breakdown)
  Sheet 2: Vật liệu        (Materials)
  Sheet 3: Nhân công        (Labor — TT 12/2021/TT-BXD rates)
  Sheet 4: Máy thi công    (Equipment)
  Sheet 5: Theo hạng mục   (By section — móng/cột/dầm/sàn/tường/hoàn thiện)
  Sheet 6: Đơn giá tổng hợp (Combined unit rates)

Formatting:
  - Headers: bold + bg color #FDE68A (Tailwind amber-200)
  - Numbers: VND currency format with thousands separators
  - Bottom totals row: bold + bg color #FCD34D (amber-300)
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any

log = logging.getLogger("zeni.boq_excel")


# ─── Style constants ────────────────────────────────────────────
HEADER_FILL = "FDE68A"  # amber-200
TOTALS_FILL = "FCD34D"  # amber-300
SECTION_FILL = "FEF3C7"  # amber-100
THIN_BORDER_STYLE = "thin"


def _apply_header_style(ws, row: int, n_cols: int) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    font = Font(bold=True, size=11)
    border = Border(
        top=Side(style=THIN_BORDER_STYLE),
        bottom=Side(style=THIN_BORDER_STYLE),
        left=Side(style=THIN_BORDER_STYLE),
        right=Side(style=THIN_BORDER_STYLE),
    )
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = align


def _apply_totals_style(ws, row: int, n_cols: int) -> None:
    from openpyxl.styles import Border, Font, PatternFill, Side

    fill = PatternFill(start_color=TOTALS_FILL, end_color=TOTALS_FILL, fill_type="solid")
    font = Font(bold=True, size=11)
    border = Border(
        top=Side(style=THIN_BORDER_STYLE),
        bottom=Side(style="double"),
        left=Side(style=THIN_BORDER_STYLE),
        right=Side(style=THIN_BORDER_STYLE),
    )
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border


def _set_vnd_format(ws, row: int, col: int) -> None:
    cell = ws.cell(row=row, column=col)
    cell.number_format = '#,##0" ₫"'


def _autosize(ws, widths: list[float]) -> None:
    """Set column widths (use letters A,B,C,...)."""
    for i, w in enumerate(widths):
        col_letter = chr(ord("A") + i)
        ws.column_dimensions[col_letter].width = w


# ───────────────────────────────────────────────────────────────
# Sheet builders
# ───────────────────────────────────────────────────────────────
def _build_sheet_summary(ws, boq_result: dict[str, Any], project_name: str, location: str) -> None:
    """Sheet 1: Tổng hợp."""
    from openpyxl.styles import Alignment, Font

    summary = boq_result.get("summary", {})
    total_vnd = int(summary.get("total_vnd", 0) or 0)
    per_m2 = int(summary.get("per_m2_vnd", 0) or 0)
    breakdown = summary.get("breakdown", {})

    ws.title = "Tổng hợp"
    ws["A1"] = "BẢNG TỔNG HỢP DỰ TOÁN XÂY DỰNG"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Dự án: {project_name}"
    ws["A2"].font = Font(bold=True, size=11)
    ws.merge_cells("A2:E2")
    ws["A3"] = f"Địa điểm: {location}"
    ws.merge_cells("A3:E3")
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    ws["A4"] = f"Ngày lập: {now}  |  Theo: TT 13/2021/TT-BXD"
    ws.merge_cells("A4:E4")

    # Header
    headers = ["STT", "Hạng mục", "Tỉ trọng (%)", "Thành tiền (VND)", "Ghi chú"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=6, column=i, value=h)
    _apply_header_style(ws, 6, len(headers))

    label_vi = {
        "phan_tho": "Phần thô (móng/khung/sàn/tường)",
        "hoan_thien": "Hoàn thiện (sơn/gạch/cửa)",
        "dien": "Hệ thống điện",
        "nuoc": "Hệ thống nước",
        "noi_that": "Nội thất",
        "khac": "Chi phí khác (10% dự phòng)",
    }
    row = 7
    for stt, (k, lbl) in enumerate(label_vi.items(), start=1):
        ratio = float(breakdown.get(k, 0) or 0)
        amount = int(total_vnd * ratio)
        ws.cell(row=row, column=1, value=stt)
        ws.cell(row=row, column=2, value=lbl)
        ws.cell(row=row, column=3, value=round(ratio * 100, 1))
        ws.cell(row=row, column=4, value=amount)
        _set_vnd_format(ws, row, 4)
        ws.cell(row=row, column=5, value="")
        row += 1

    # Totals row
    ws.cell(row=row, column=2, value="TỔNG CỘNG")
    ws.cell(row=row, column=3, value=100.0)
    ws.cell(row=row, column=4, value=total_vnd)
    _set_vnd_format(ws, row, 4)
    _apply_totals_style(ws, row, len(headers))

    # Per-m² info
    ws.cell(row=row + 2, column=1, value=f"Đơn giá m² xây dựng: {per_m2:,} ₫/m²")
    ws.cell(row=row + 2, column=1).font = Font(bold=True, italic=True)

    _autosize(ws, [6, 42, 14, 22, 32])


def _build_sheet_materials(ws, boq_result: dict[str, Any]) -> None:
    """Sheet 2: Vật liệu."""
    from openpyxl.styles import Font

    ws.title = "Vật liệu"
    headers = ["STT", "Mã VT", "Tên vật tư", "Đơn vị", "Khối lượng", "Đơn giá (VND)", "Thành tiền (VND)", "Nguồn cung"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_style(ws, 1, len(headers))

    materials = boq_result.get("materials") or _default_materials()
    total = 0
    for stt, item in enumerate(materials, start=1):
        qty = float(item.get("quantity", 0) or 0)
        unit_price = int(item.get("unit_price_vnd", 0) or 0)
        amount = int(qty * unit_price)
        total += amount
        r = stt + 1
        ws.cell(row=r, column=1, value=stt)
        ws.cell(row=r, column=2, value=item.get("code", f"VT.{stt:03d}"))
        ws.cell(row=r, column=3, value=item.get("name", "—"))
        ws.cell(row=r, column=4, value=item.get("unit", "kg"))
        ws.cell(row=r, column=5, value=qty)
        ws.cell(row=r, column=6, value=unit_price)
        _set_vnd_format(ws, r, 6)
        ws.cell(row=r, column=7, value=amount)
        _set_vnd_format(ws, r, 7)
        ws.cell(row=r, column=8, value=item.get("source", ""))

    total_row = len(materials) + 2
    ws.cell(row=total_row, column=3, value="TỔNG VẬT LIỆU").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total)
    _set_vnd_format(ws, total_row, 7)
    _apply_totals_style(ws, total_row, len(headers))

    _autosize(ws, [6, 10, 36, 8, 12, 18, 20, 22])


def _build_sheet_labor(ws, boq_result: dict[str, Any]) -> None:
    """Sheet 3: Nhân công — TT 12/2021/TT-BXD rates."""
    from openpyxl.styles import Font

    ws.title = "Nhân công"
    headers = ["STT", "Mã", "Bậc thợ / chức danh", "Đơn vị", "Số công", "Đơn giá (VND/công)", "Thành tiền (VND)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_style(ws, 1, len(headers))
    ws.cell(row=2, column=1,
            value="Tham chiếu: TT 12/2021/TT-BXD — Đơn giá nhân công xây dựng").font = Font(italic=True)

    labor = boq_result.get("labor") or _default_labor()
    total = 0
    for stt, item in enumerate(labor, start=1):
        n_days = float(item.get("man_days", 0) or 0)
        rate = int(item.get("rate_per_day_vnd", 0) or 0)
        amount = int(n_days * rate)
        total += amount
        r = stt + 3
        ws.cell(row=r, column=1, value=stt)
        ws.cell(row=r, column=2, value=item.get("code", f"NC.{stt:03d}"))
        ws.cell(row=r, column=3, value=item.get("role", "—"))
        ws.cell(row=r, column=4, value="công")
        ws.cell(row=r, column=5, value=n_days)
        ws.cell(row=r, column=6, value=rate)
        _set_vnd_format(ws, r, 6)
        ws.cell(row=r, column=7, value=amount)
        _set_vnd_format(ws, r, 7)

    total_row = len(labor) + 4
    ws.cell(row=total_row, column=3, value="TỔNG NHÂN CÔNG").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total)
    _set_vnd_format(ws, total_row, 7)
    _apply_totals_style(ws, total_row, len(headers))

    _autosize(ws, [6, 10, 36, 8, 12, 22, 22])


def _build_sheet_equipment(ws, boq_result: dict[str, Any]) -> None:
    """Sheet 4: Máy thi công — TT 13/2021/TT-BXD."""
    from openpyxl.styles import Font

    ws.title = "Máy thi công"
    headers = ["STT", "Mã máy", "Tên máy", "Đơn vị", "Số ca", "Đơn giá ca (VND)", "Thành tiền (VND)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_style(ws, 1, len(headers))
    ws.cell(row=2, column=1,
            value="Tham chiếu: TT 13/2021/TT-BXD — Định mức ca máy thi công").font = Font(italic=True)

    equipment = boq_result.get("equipment") or _default_equipment()
    total = 0
    for stt, item in enumerate(equipment, start=1):
        ca = float(item.get("shifts", 0) or 0)
        rate = int(item.get("rate_per_shift_vnd", 0) or 0)
        amount = int(ca * rate)
        total += amount
        r = stt + 3
        ws.cell(row=r, column=1, value=stt)
        ws.cell(row=r, column=2, value=item.get("code", f"MTC.{stt:03d}"))
        ws.cell(row=r, column=3, value=item.get("name", "—"))
        ws.cell(row=r, column=4, value="ca")
        ws.cell(row=r, column=5, value=ca)
        ws.cell(row=r, column=6, value=rate)
        _set_vnd_format(ws, r, 6)
        ws.cell(row=r, column=7, value=amount)
        _set_vnd_format(ws, r, 7)

    total_row = len(equipment) + 4
    ws.cell(row=total_row, column=3, value="TỔNG MÁY THI CÔNG").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total)
    _set_vnd_format(ws, total_row, 7)
    _apply_totals_style(ws, total_row, len(headers))

    _autosize(ws, [6, 12, 36, 8, 12, 22, 22])


def _build_sheet_by_section(ws, boq_result: dict[str, Any]) -> None:
    """Sheet 5: Theo hạng mục."""
    from openpyxl.styles import Font, PatternFill

    ws.title = "Theo hạng mục"
    headers = ["STT", "Hạng mục", "Đơn vị", "Khối lượng", "Đơn giá (VND)", "Thành tiền (VND)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_style(ws, 1, len(headers))

    sections = boq_result.get("by_section") or _default_by_section()
    section_fill = PatternFill(start_color=SECTION_FILL, end_color=SECTION_FILL, fill_type="solid")

    row = 2
    grand_total = 0
    for section_name, items in sections.items():
        ws.cell(row=row, column=2, value=section_name).font = Font(bold=True, size=11)
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = section_fill
        row += 1
        section_total = 0
        for stt, item in enumerate(items, start=1):
            qty = float(item.get("quantity", 0) or 0)
            unit_price = int(item.get("unit_price_vnd", 0) or 0)
            amount = int(qty * unit_price)
            section_total += amount
            ws.cell(row=row, column=1, value=stt)
            ws.cell(row=row, column=2, value=item.get("name", "—"))
            ws.cell(row=row, column=3, value=item.get("unit", "—"))
            ws.cell(row=row, column=4, value=qty)
            ws.cell(row=row, column=5, value=unit_price)
            _set_vnd_format(ws, row, 5)
            ws.cell(row=row, column=6, value=amount)
            _set_vnd_format(ws, row, 6)
            row += 1
        ws.cell(row=row, column=2, value=f"Cộng {section_name}").font = Font(bold=True, italic=True)
        ws.cell(row=row, column=6, value=section_total)
        _set_vnd_format(ws, row, 6)
        grand_total += section_total
        row += 2

    ws.cell(row=row, column=2, value="TỔNG CỘNG (TẤT CẢ HẠNG MỤC)")
    ws.cell(row=row, column=6, value=grand_total)
    _set_vnd_format(ws, row, 6)
    _apply_totals_style(ws, row, len(headers))

    _autosize(ws, [6, 40, 10, 14, 20, 22])


def _build_sheet_combined_rates(ws, boq_result: dict[str, Any]) -> None:
    """Sheet 6: Đơn giá tổng hợp."""
    from openpyxl.styles import Font

    ws.title = "Đơn giá tổng hợp"
    headers = ["STT", "Mã hiệu", "Hạng mục công tác", "Đơn vị",
               "VL (VND)", "NC (VND)", "M (VND)", "Tổng (VND)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_style(ws, 1, len(headers))

    rates = boq_result.get("combined_rates") or _default_combined_rates()
    for stt, item in enumerate(rates, start=1):
        vl = int(item.get("material_vnd", 0) or 0)
        nc = int(item.get("labor_vnd", 0) or 0)
        m = int(item.get("equipment_vnd", 0) or 0)
        total = vl + nc + m
        r = stt + 1
        ws.cell(row=r, column=1, value=stt)
        ws.cell(row=r, column=2, value=item.get("code", f"AF.{stt:03d}"))
        ws.cell(row=r, column=3, value=item.get("name", "—"))
        ws.cell(row=r, column=4, value=item.get("unit", "—"))
        ws.cell(row=r, column=5, value=vl)
        _set_vnd_format(ws, r, 5)
        ws.cell(row=r, column=6, value=nc)
        _set_vnd_format(ws, r, 6)
        ws.cell(row=r, column=7, value=m)
        _set_vnd_format(ws, r, 7)
        ws.cell(row=r, column=8, value=total)
        _set_vnd_format(ws, r, 8)

    note_row = len(rates) + 3
    ws.cell(row=note_row, column=1,
            value="Đơn giá tổng hợp = VL (Vật liệu) + NC (Nhân công) + M (Máy)").font = Font(italic=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)

    _autosize(ws, [6, 12, 40, 10, 16, 16, 16, 20])


# ───────────────────────────────────────────────────────────────
# Default fallback data when boq_result lacks fields
# ───────────────────────────────────────────────────────────────
def _default_materials() -> list[dict[str, Any]]:
    return [
        {"code": "VT.001", "name": "Xi măng PCB30 Hà Tiên (bao 50kg)", "unit": "bao", "quantity": 200, "unit_price_vnd": 95000, "source": "Hà Tiên"},
        {"code": "VT.002", "name": "Cát vàng đổ bê tông", "unit": "m³", "quantity": 18, "unit_price_vnd": 480000, "source": "Bình Dương"},
        {"code": "VT.003", "name": "Đá 1x2 dăm sàng", "unit": "m³", "quantity": 22, "unit_price_vnd": 420000, "source": "Đồng Nai"},
        {"code": "VT.004", "name": "Thép Việt-Nhật d10-d22", "unit": "kg", "quantity": 4500, "unit_price_vnd": 22000, "source": "Việt-Nhật"},
        {"code": "VT.005", "name": "Gạch tuynel 6 lỗ Đại Thanh", "unit": "viên", "quantity": 8000, "unit_price_vnd": 2200, "source": "Đại Thanh"},
        {"code": "VT.006", "name": "Sơn nội thất Dulux Inspire trắng", "unit": "thùng", "quantity": 12, "unit_price_vnd": 1850000, "source": "AkzoNobel"},
        {"code": "VT.007", "name": "Gạch lát Đồng Tâm 600x600", "unit": "m²", "quantity": 180, "unit_price_vnd": 280000, "source": "Đồng Tâm"},
    ]


def _default_labor() -> list[dict[str, Any]]:
    return [
        {"code": "NC.001", "role": "Thợ chính (bậc 5/7)", "man_days": 120, "rate_per_day_vnd": 550000},
        {"code": "NC.002", "role": "Thợ phụ (bậc 3/7)",   "man_days": 180, "rate_per_day_vnd": 380000},
        {"code": "NC.003", "role": "Thợ điện (bậc 4/7)",  "man_days": 30, "rate_per_day_vnd": 480000},
        {"code": "NC.004", "role": "Thợ nước (bậc 4/7)",  "man_days": 25, "rate_per_day_vnd": 470000},
        {"code": "NC.005", "role": "Thợ sơn (bậc 4/7)",   "man_days": 22, "rate_per_day_vnd": 450000},
    ]


def _default_equipment() -> list[dict[str, Any]]:
    return [
        {"code": "MTC.001", "name": "Máy trộn bê tông 250L",     "shifts": 12, "rate_per_shift_vnd": 480000},
        {"code": "MTC.002", "name": "Máy đầm dùi bê tông 1.5kW", "shifts": 10, "rate_per_shift_vnd": 220000},
        {"code": "MTC.003", "name": "Máy cắt sắt thanh φ32",     "shifts": 8,  "rate_per_shift_vnd": 320000},
        {"code": "MTC.004", "name": "Máy hàn điện 23kW",          "shifts": 6,  "rate_per_shift_vnd": 380000},
        {"code": "MTC.005", "name": "Cẩu tự hành 5T",             "shifts": 4,  "rate_per_shift_vnd": 2200000},
    ]


def _default_by_section() -> dict[str, list[dict[str, Any]]]:
    return {
        "1. Móng": [
            {"name": "Đào đất móng (đất cấp 2)", "unit": "m³", "quantity": 18, "unit_price_vnd": 165000},
            {"name": "Bê tông lót móng đá 4x6",   "unit": "m³", "quantity": 3.2, "unit_price_vnd": 920000},
            {"name": "Bê tông móng M250",         "unit": "m³", "quantity": 8.5, "unit_price_vnd": 1350000},
        ],
        "2. Cột": [
            {"name": "Cốt thép cột (kg)", "unit": "kg",  "quantity": 1200, "unit_price_vnd": 24500},
            {"name": "Bê tông cột M250",  "unit": "m³",  "quantity": 4.2,  "unit_price_vnd": 1400000},
            {"name": "Ván khuôn cột",     "unit": "m²",  "quantity": 56,   "unit_price_vnd": 185000},
        ],
        "3. Dầm": [
            {"name": "Cốt thép dầm (kg)", "unit": "kg",  "quantity": 1800, "unit_price_vnd": 24500},
            {"name": "Bê tông dầm M250",  "unit": "m³",  "quantity": 5.8,  "unit_price_vnd": 1380000},
        ],
        "4. Sàn": [
            {"name": "Cốt thép sàn (kg)", "unit": "kg",  "quantity": 2400, "unit_price_vnd": 23800},
            {"name": "Bê tông sàn M250",  "unit": "m³",  "quantity": 9.6,  "unit_price_vnd": 1360000},
        ],
        "5. Tường": [
            {"name": "Xây tường gạch tuynel 200mm", "unit": "m²", "quantity": 220, "unit_price_vnd": 285000},
            {"name": "Trát tường 2 mặt vữa M75",    "unit": "m²", "quantity": 440, "unit_price_vnd": 92000},
        ],
        "6. Hoàn thiện": [
            {"name": "Sơn nội thất 2 nước",   "unit": "m²", "quantity": 380, "unit_price_vnd": 65000},
            {"name": "Lát gạch nền 600x600",  "unit": "m²", "quantity": 180, "unit_price_vnd": 320000},
            {"name": "Cửa gỗ công nghiệp 0.9x2.2m", "unit": "bộ", "quantity": 8, "unit_price_vnd": 4500000},
        ],
    }


def _default_combined_rates() -> list[dict[str, Any]]:
    return [
        {"code": "AF.001", "name": "Bê tông móng M250", "unit": "m³",
         "material_vnd": 980000, "labor_vnd": 280000, "equipment_vnd": 90000},
        {"code": "AF.002", "name": "Bê tông cột M250", "unit": "m³",
         "material_vnd": 980000, "labor_vnd": 330000, "equipment_vnd": 90000},
        {"code": "AF.003", "name": "Xây tường gạch tuynel 200mm", "unit": "m²",
         "material_vnd": 165000, "labor_vnd": 95000, "equipment_vnd": 25000},
        {"code": "AF.004", "name": "Trát tường vữa M75", "unit": "m²",
         "material_vnd": 38000, "labor_vnd": 42000, "equipment_vnd": 12000},
        {"code": "AF.005", "name": "Sơn nội thất 2 nước", "unit": "m²",
         "material_vnd": 28000, "labor_vnd": 25000, "equipment_vnd": 12000},
        {"code": "AF.006", "name": "Lát gạch nền 600x600", "unit": "m²",
         "material_vnd": 220000, "labor_vnd": 65000, "equipment_vnd": 35000},
    ]


# ───────────────────────────────────────────────────────────────
# Main entry
# ───────────────────────────────────────────────────────────────
def generate_boq_workbook(
    boq_result: dict[str, Any],
    project_name: str,
    location: str,
) -> bytes:
    """
    Generate full 6-sheet BOQ Excel workbook.

    Args:
        boq_result: dict — typically the .output of BOQCalculatorAgent.
                    Optional keys: summary, materials, labor, equipment,
                    by_section, combined_rates. Defaults used when missing.
        project_name: project name (for header)
        location: tỉnh/thành

    Returns:
        XLSX bytes — openpyxl-generated, ready to upload to GCS.
    """
    from openpyxl import Workbook  # type: ignore

    log.info("[boq_excel] generating workbook for project='%s' location='%s'",
             project_name[:40], location[:20])

    wb = Workbook()
    # Sheet 1 (default)
    _build_sheet_summary(wb.active, boq_result, project_name, location)
    # Sheets 2-6
    _build_sheet_materials(wb.create_sheet("Vật liệu"), boq_result)
    _build_sheet_labor(wb.create_sheet("Nhân công"), boq_result)
    _build_sheet_equipment(wb.create_sheet("Máy thi công"), boq_result)
    _build_sheet_by_section(wb.create_sheet("Theo hạng mục"), boq_result)
    _build_sheet_combined_rates(wb.create_sheet("Đơn giá tổng hợp"), boq_result)

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    log.info("[boq_excel] workbook generated, %d bytes", len(data))
    return data


__all__ = ["generate_boq_workbook"]
# end of file
